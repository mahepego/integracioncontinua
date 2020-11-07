from django.shortcuts import render

from django.views.generic.list import ListView
from django.views.generic.detail import DetailView
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.views.generic import TemplateView
from django.urls import reverse_lazy

from evento.models import Evento
from openpyxl import Workbook
from django.http.response import HttpResponse

# Create your views here.

class HomeView(ListView):
    model = Evento
    template_name = 'evento/index.html'
    paginate_by = 10


class EventoDetailView(DetailView):
    model = Evento
    template_name = 'evento/evento_detail.html'


class EventoCreateView(CreateView):
    model = Evento
    template_name = 'evento/evento_create.html'
    fields = ['codigo', 'nombre', 'tipo','duracion','descripcion']


class EventoUpdateView(UpdateView):
    model = Evento
    template_name = 'evento/evento_create.html'
    fields = ['codigo', 'nombre', 'tipo','duracion','descripcion']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['edit'] = True
        return context


class EventoDeleteView(DeleteView):
    model = Evento
    success_url = reverse_lazy('evento:homeEvento')
    template_name = 'evento/confirm_evento_deletion.html'

class ConsolidadoEvento(TemplateView):
    def get(self,request,*args,**kwargs):
        eventos = Evento.objects.all()
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'CONSOLIDADO DE EVENTOS'
        ws.merge_cells('B1:F1')
        ws['B3'] = 'CODIGO'
        ws['C3'] = 'NOMBRE'
        ws['D3'] = 'TIPO'
        ws['E3'] = 'DURACION'
        ws['F3'] = 'DESCRIPCION'
        cont = 4

        for evento in eventos:
            ws.cell(row = cont, column = 2).value = evento.codigo
            ws.cell(row = cont, column = 3).value = evento.nombre
            ws.cell(row = cont, column = 4).value = evento.tipo
            ws.cell(row = cont, column = 5).value = evento.duracion
            ws.cell(row = cont, column = 6).value = evento.descripcion
            cont+=1

        nombre_archivo = "ConsolidadoEvento.xlsx"
        response = HttpResponse(content_type = "applicacion/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response