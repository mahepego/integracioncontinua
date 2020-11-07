from django.shortcuts import render

from django.views.generic.list import ListView
from django.views.generic.detail import DetailView
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy

from django.views.generic import TemplateView
from persona.models import Persona

from openpyxl import Workbook
from django.http.response import HttpResponse

# Create your views here.

class HomeView(ListView):
    model = Persona
    template_name = 'persona/index.html'
    paginate_by = 10


class PersonaDetailView(DetailView):
    model = Persona
    template_name = 'persona/persona_detail.html'


class PersonaCreateView(CreateView):
    model = Persona
    template_name = 'persona/persona_create.html'
    fields = ['numIdentificacion', 'tipoIdentificacion', 'apellidoPaterno','apellidoMaterno','nombres','fechaNacimiento','sexo','telefono','direccion','email','evento']


class PersonaUpdateView(UpdateView):
    model = Persona
    template_name = 'persona/persona_create.html'
    fields = ['numIdentificacion', 'tipoIdentificacion', 'apellidoPaterno','apellidoMaterno','nombres','fechaNacimiento','sexo','telefono','direccion','email','evento']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['edit'] = True
        return context


class PersonaDeleteView(DeleteView):
    model = Persona
    success_url = reverse_lazy('persona:homePersona')
    template_name = 'persona/confirm_persona_deletion.html'

class ConsolidadoPersona(TemplateView):
    def get(self,request,*args,**kwargs):
        personas = Persona.objects.all()
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'CONSOLIDADO DE PERSONAS'
        ws.merge_cells('B1:L1')
        ws['B3'] = 'NO IDENTIFICACION'
        ws['C3'] = 'TIPO IDENTIFICACION'
        ws['D3'] = 'APELLIDO PATERNO'
        ws['E3'] = 'APELLIDO MATERNO'
        ws['F3'] = 'NOMBRES'
        ws['G3'] = 'FECHA NACIMIENTO'
        ws['H3'] = 'SEXO'
        ws['I3'] = 'TELEFONO'
        ws['J3'] = 'DIRECCION'
        ws['K3'] = 'EMAIL'
        #ws['L3'] = 'EVENTO'
        cont = 4

        for persona in personas:
            ws.cell(row = cont, column = 2).value = persona.numIdentificacion
            ws.cell(row = cont, column = 3).value = persona.tipoIdentificacion
            ws.cell(row = cont, column = 4).value = persona.apellidoPaterno
            ws.cell(row = cont, column = 5).value = persona.apellidoMaterno
            ws.cell(row = cont, column = 6).value = persona.nombres
            ws.cell(row = cont, column = 7).value = persona.fechaNacimiento
            ws.cell(row = cont, column = 8).value = persona.sexo
            ws.cell(row = cont, column = 9).value = persona.telefono
            ws.cell(row = cont, column = 10).value = persona.direccion
            ws.cell(row = cont, column = 11).value = persona.email
            #ws.cell(row = cont, column = 12).value = persona.evento
            cont+=1

        nombre_archivo = "ConsolidadoPersona.xlsx"
        response = HttpResponse(content_type = "applicacion/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response