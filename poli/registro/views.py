from django.shortcuts import render

from django.views.generic.list import ListView
from django.views.generic.detail import DetailView
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.views.generic import TemplateView

from registro.models import Registro

from openpyxl import Workbook
from django.http.response import HttpResponse

# Create your views here.

class HomeView(ListView):
    model = Registro
    template_name = 'registro/index.html'
    paginate_by = 10


class RegistroDetailView(DetailView):
    model = Registro
    template_name = 'registro/registro_detail.html'


class RegistroCreateView(CreateView):
    model = Registro
    template_name = 'registro/registro_create.html'
    fields = ['persona', 'evento']


class RegistroUpdateView(UpdateView):
    model = Registro
    template_name = 'registro/registro_create.html'
    fields = ['persona', 'evento']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['edit'] = True
        return context


class RegistroDeleteView(DeleteView):
    model = Registro
    success_url = reverse_lazy('registro:homeRegistro')
    template_name = 'registro/confirm_registro_deletion.html'

class ConsolidadoRegistro(TemplateView):
    def get(self,request,*args,**kwargs):
        registros = Registro.objects.all()
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'CONSOLIDADO DE REGISTROS'
        ws.merge_cells('B1:D1')
        #ws['B3'] = 'PERSONA'
        #ws['C3'] = 'EVENTO'
        ws['D3'] = 'FECHA DE EVENTO'
        cont = 4

        for registro in registros:
         #   ws.cell(row = cont, column = 2).value = registro.persona
         #   ws.cell(row = cont, column = 3).value = registro.evento
            ws.cell(row = cont, column = 4).value = registro.fechaEvento
            cont+=1

        nombre_archivo = "ConsolidadoRegistros.xlsx"
        response = HttpResponse(content_type = "applicacion/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response