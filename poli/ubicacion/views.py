from django.shortcuts import render

from django.views.generic.list import ListView
from django.views.generic.detail import DetailView
from django.views.generic.edit import CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.views.generic import TemplateView

from ubicacion.models import Ubicacion

from openpyxl import Workbook
from django.http.response import HttpResponse

# Create your views here.

class HomeView(ListView):
    model = Ubicacion
    template_name = 'ubicacion/index.html'
    paginate_by = 10


class UbicacionDetailView(DetailView):
    model = Ubicacion
    template_name = 'ubicacion/ubicacion_detail.html'


class UbicacionCreateView(CreateView):
    model = Ubicacion
    template_name = 'ubicacion/ubicacion_create.html'
    fields = ['codigo', 'sector', 'nombre', 'capacidad']


class UbicacionUpdateView(UpdateView):
    model = Ubicacion
    template_name = 'ubicacion/ubicacion_create.html'
    fields = ['codigo', 'sector', 'nombre', 'capacidad']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['edit'] = True
        return context


class UbicacionDeleteView(DeleteView):
    model = Ubicacion
    success_url = reverse_lazy('ubicacion:homeUbicacion')
    template_name = 'ubicacion/confirm_ubicacion_deletion.html'

class ConsolidadoUbicacion(TemplateView):
    def get(self,request,*args,**kwargs):
        ubicaciones = Ubicacion.objects.all()
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'CONSOLIDADO DE UBICACIONES'
        ws.merge_cells('B1:E1')
        ws['B3'] = 'CODIGO'
        ws['C3'] = 'SECTOR'
        ws['D3'] = 'NOMBRE'
        ws['E3'] = 'CAPACIDAD'
        cont = 4

        for ubicacion in ubicaciones:
            ws.cell(row = cont, column = 2).value = ubicacion.codigo
            ws.cell(row = cont, column = 3).value = ubicacion.sector
            ws.cell(row = cont, column = 4).value = ubicacion.nombre
            ws.cell(row = cont, column = 5).value = ubicacion.capacidad
            cont+=1

        nombre_archivo = "ConsolidadoUbicacion.xlsx"
        response = HttpResponse(content_type = "applicacion/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response